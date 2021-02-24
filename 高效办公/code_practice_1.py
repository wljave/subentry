from openpyxl import load_workbook

# 打开【practice_final.xlsx】工作簿
staff_wb = load_workbook('./practice0.xlsx')
# 获取【上半年公司名单】工作表
active_ws = staff_wb['上半年公司名单']

# 获取指定位置单元格的值，并打印
rows = active_ws.iter_rows(min_row=11, max_row=13, min_col=2, max_col=3, values_only=True)
for row in rows:
    print(row)

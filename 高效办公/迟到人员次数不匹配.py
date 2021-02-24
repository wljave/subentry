from openpyxl import load_workbook

# 打开工作簿【10月考勤统计.xlsx】，获取活动工作表
wb = load_workbook('./10月考勤统计.xlsx')
ws = wb.active

# 创建迟到人员字典
info_dict = {}

# 循环读取除表头外的表格数据
for row in ws.iter_rows(min_row=2, values_only=True):
    # 取出员工工号
    staff_id = row[0]
    # 取出迟到次数
    staff_late = row[-1]
    # 将信息添加入字典，字典格式为{'员工工号': '迟到次数'}
    info_dict[staff_id] = staff_late

# 打开工作簿【迟到次数月度统计（10月更新）.xlsx】，获取活动工作表
monthly_wb = load_workbook('./迟到次数月度统计（10月更新）.xlsx')
monthly_ws = monthly_wb.active

# 循环读取出表头外的表格数据
for monthly_row in monthly_ws.iter_rows(min_row=3, max_col=13, values_only=True):
    # 取出员工工号
    member_id = monthly_row[0]
    # 取出十月份的迟到次数
    member_late = monthly_row[-1]
    # 匹配迟到次数是否相等
    if member_late != info_dict[member_id]:
        print('工号{}迟到情况不匹配，请核查后更新'.format(member_id))

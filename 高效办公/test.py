from openpyxl import load_workbook

wb = load_workbook('./10月考勤统计.xlsx')
ws = wb['Sheet1']

list1 = ['吕建国', '车敏', '张燕', '刘晨']
dict1 = {}

for row in ws.iter_rows(min_row=2, min_col=2, max_col=3, values_only=True):
    dict1[row[0]] = row[1]
    if row[0] in list1:
        print(dict1[row[0]])


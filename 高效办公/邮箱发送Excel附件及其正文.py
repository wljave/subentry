from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# 设置邮箱账号
# account = input('请输入邮箱账户：')
account = '540247580@qq.com'
# 设置邮箱授权码
# token = input('请输入邮箱授权码：')
token = 'bbtyakjogxdybdda'
# 设置邮箱服务器，端口
smtp = smtplib.SMTP_SSL('smtp.qq.com', 465)
# 登录qq邮箱
smtp.login(account, token)

# 打开工作表
wb = load_workbook('./10月考勤统计.xlsx')
sheet = wb.active

# 编写正文内容
content = '10月的考勤表已出，其中迟到时长超出 45 分钟的人员如下：\n'
for row_data in sheet.iter_rows(min_row=2, values_only=True):
    # 获取迟到时长超过45分钟的人员
    if row_data[3] > 45:
        content += '姓名：{name} 迟到总时长：{time} 迟到次数：{degree}\n'.format(name=row_data[1], time=row_data[3], degree=row_data[4])
content += '详情见附件内容'

# 创建简单邮件对象
email_content = MIMEText(content, 'plain', 'utf-8')

# 读取工作表文件数据
with open('./10月考勤统计.xlsx', 'rb') as f:
    file_data = f.read()
# 设置内容类型为附件
attachment = MIMEText(file_data, 'base64', 'utf-8')
# 设置附件标题以及文件类型
attachment.add_header('Content-Disposition', 'attachment', filename='10月考勤统计.xlsx')

# 创建复合邮件对象
msg = MIMEMultipart()

# 添加正文到复合邮件对象中
msg.attach(email_content)

# 添加附件到复合邮件对象里
msg.attach(attachment)

# 设置发送者信息
msg['From'] = '陈知枫'
# 设置接受者信息
msg['To'] = '闪光金融的各位同事们'
# 设置邮件标题
msg['Subject'] = '10月考勤统计'

try:
    # 发送邮件
    # 三个参数分别是：发件人邮箱账号，收件人邮箱账号，发送的邮件体
    smtp.sendmail(account, 'xiaofei2025@163.com', msg.as_string())
    print('电子邮件发送成功!')
except:
    print('邮件发送失败，请重试！')

# 退出服务器，结束SMTP会话
# 关闭邮箱服务
smtp.quit()

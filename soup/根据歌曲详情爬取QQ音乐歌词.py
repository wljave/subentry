import requests
import openpyxl
import csv
import xlsxwriter
# from bs4 import BeautifulSoup
# from urllib.request import quote


# 定义请求头部信息
def reptile_header(url, params):
    # 模拟浏览器请求数据
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/87.0.4280.88 Safari/537.36 ',
        # 爬取QQ音乐里面的歌词信息的时候，如果不添加origin和referer就会读取不到数据【{'retcode': -1310, 'code': -1310, 'subcode': -1310}】
        'origin': 'https://y.qq.com',
        'referer': 'https://y.qq.com/'
    }
    # 设置代理IP
    proxy = '103.109.59.242:53281'
    # 需要认证的代理可以写成如下格式，username为用户名，password为密码
    # proxy = "username:password@ip:端口号"
    proxies = {
        'http': 'https://' + proxy,
    }
    res = requests.get(url=url, params=params, proxies=proxies, headers=headers)
    # 判断状态码是否正常
    # print(res.status_code)
    return res


# 定义ID列
ids_column = []
# 定义歌曲名列
song_name_column = []
# 定义歌词列表
song_lyric = []
# 定义初始化评论总条数
comment_totals = []
# 定义不存在的歌手名
add_file = []
# 提示详情信息及链接
song_name_keyword = input('请输入要搜索的歌手名字(爬取歌词)：')


# search_url = 'https://y.qq.com/portal/search.html'
# parm = {
#     'page': '1',
#     'searchid': '1',
#     'remoteplace': 'txt.yqq.top',
#     't': 'song',
#     'w': song_name_keyword
# }
#
# res_music = reptile_header(url=search_url, params=parm)
# res_music.encoding = 'utf-8'
# # 判断状态码是否正常
# if res_music.status_code == 200:
#     print('搜索请求正常')
# print('搜索结果链接：' + res_music.url)


# # 获得歌词id，为爬取歌词最准备
# list_music_lyric_ids = []
# list_music_name = []
#
#
# def song_message(param):
#     global list_music_lyric_ids, list_music_name
#     xhr = 'https://c.y.qq.com/soso/fcgi-bin/client_search_cp'
#     # 调用get方法，下载这个字典
#     res_music_message = reptile_header(url=xhr, params=param)
#     res_music_message.encoding = 'utf-8'
#     # 判断连接是否正常
#     if res_music_message.status_code == 200:
#         print('连接正常，可以获取歌词id\n')
#     # 使用json()方法，将response对象，转为列表/字典
#     json_music = res_music_message.json()
#     # 一层一层地取字典，获取歌单列表
#     list_music = json_music['data']['song']['list']
#     # list_music是一个列表，music是它里面的元素
#
#     for music in list_music:
#         # 获得歌词id，为爬取歌词最准备
#         music_lyric_id = music['id']
#         list_music_lyric_ids.append(music_lyric_id)
#         # print('序号：' + str(music_lyric_id))
#         # 以name为键，查找歌曲名
#         music_name = music['name']
#         list_music_name.append(music_name)
#
#     print(list_music_lyric_ids)
#     print(list_music_name)
#
#     return list_music_lyric_ids, list_music_name
#
#
# p = 1
# while p < 3:
#     # 每页显示的数量最大显示60条
#     param = {
#         'ct': '24',
#         'qqmusic_ver': '1298',
#         'new_json': '1',
#         'remoteplace': 'txt.yqq.song',
#         'searchid': '67751296646578192',
#         't': '0',
#         'aggr': '1',
#         'cr': '1',
#         'catZhida': '1',
#         'lossless': '0',
#         'flag_qc': '0',
#         'p': p,
#         'n': '60',
#         'w': song_name_keyword,
#         'g_tk_new_20200303': '1464878100',
#         'g_tk': '1592435056',
#         'loginUin': '0',
#         'hostUin': '0',
#         'format': 'json',
#         'inCharset': 'utf8',
#         'outCharset': 'utf-8',
#         'notice': '0',
#         'platform': 'yqq.json',
#         'needNewCode': '0'
#     }
#     p += 1
#     song_message(param)


# 爬取歌词
# ******读取excel表中的歌曲ID*******

try:
    # 0打开工作薄
    wb = openpyxl.load_workbook('./crawl_song/%s-歌曲详细信息.xlsx' % song_name_keyword)
    # 1获取工作簿
    # print(wb.active)    # 查看已有的工作簿，并打印
    sheet = wb[song_name_keyword]
    # sheetname = wb.sheetnames   # 将获取到的工作簿名称并赋值
    # print(sheetname)
    # 2读取A列数据
    for cell_A in sheet['A']:
        # print(cell.value)
        ids_column.append(cell_A.value)
    for cell_B in sheet['B']:
        song_name_column.append(cell_B.value)
    # print(ids_column)
    # 删除A列的表头
    del ids_column[0]
    del song_name_column[0]
    print(ids_column)
    print(song_name_column)

    # ******爬取评论总页数*******
    # 抓取所有歌曲的评论总数量comment_total，加入列表comment_totals中
    def new_comment_total(pas):
        # global comment_total
        # 定义最新评论链接
        comment_url = 'https://c.y.qq.com/base/fcgi-bin/fcg_global_comment_h5.fcg'
        # 使用json()方法，将response对象，转为列表/字典
        json_comment = reptile_header(url=comment_url, params=pas).json()
        # 查看JSON里面都有什么数据
        # print(json_comment)
        # 获取该首歌曲最新评论的总条数 comment_total 是int类型
        comment_total = json_comment['comment']['commenttotal']
        comment_totals.append(comment_total)


    # for n in range(len(ids_column)):
    #     pass

    k = 0
    # # ******创建csv表用来保存歌曲名*******
    #
    # with open('./lyric/%s-歌词.csv' % song_name_keyword, 'a+', newline='', encoding='utf-8')as initial:
    #     write_title = csv.writer(initial)
    #     write_title.writerow(['歌曲名', song_name_column[k - 1]])

    # 爬取歌词
    def lyric(par):

        lyric_url = 'https://c.y.qq.com/lyric/fcgi-bin/fcg_query_lyric_yqq.fcg'

        # 调用get方法，下载这个字典
        res_lyric = reptile_header(url=lyric_url, params=par)
        res_lyric.encoding = 'utf-8'
        if res_lyric.status_code == 200:
            print('连接正常，开始爬取歌词')

        # 爬取歌词json字典
        music_lyric = res_lyric.json()
        # ******创建Excel表用来保存歌曲名*******
        # 创建Excel文件
        nb = xlsxwriter.Workbook('./lyric/%s-歌词.xlsx' % song_name_keyword)
        # 获取工作簿的活动表
        # sheet = wb.worksheets()
        # 创建工作表
        sheets = nb.add_worksheet(song_name_keyword)
        # 添加工作表样式
        bold = nb.add_format({
            'bold': True,  # 字体加粗
            'border': 1,  # 单元格边框宽度
            'align': 'left',  # 水平对齐方式
            'valign': 'vcenter',  # 垂直对齐方式
            'fg_color': '#819FF7',  # 单元格背景颜色
            'text_wrap': True,  # 是否自动换行
        })
        # 写入一整行 row:行， col：列， data:要写入的数据, bold:单元格的样式
        sheets.write_row('A1', ['歌曲ID', '歌曲名', '歌词', '评论总数'], bold)
        # # ******创建csv表用来保存歌词*******
        # with open('./lyric/%s-歌词.csv' % song_name_keyword, 'a+', newline='', encoding='utf-8')as ly:
        #     write_lyric = csv.writer(ly)
        #     write_lyric.writerow(['歌词'])
        sheets.write_column('A2', ids_column)
        sheets.write_column('B2', song_name_column)
        sheets.write_column('D2', comment_totals)
        try:
            lyrics = music_lyric['lyric']
            song_lyric.append(lyrics)
            # 写入一整行
            sheets.write_column('C2', song_lyric)
            # # write_lyric.writerow(['歌词', lyrics])
        except KeyError:
            lyrics = ' ┝ 暂无歌词'
            song_lyric.append(lyrics)
            # 写入一整行
            # sheets.write_column('A2', ids_column)
            # sheets.write_column('B2', song_name_column)
            sheets.write_column('C2', song_lyric)
            print(lyrics)
        nb.close()

    # 根据歌词id查找对应的歌词
    for ids in ids_column:
        par = {
            'nobase64': '1',
            'musicid': ids,
            '-': 'jsonp1',
            'g_tk_new_20200303': '',
            'g_tk': '',
            'loginUin': '0',
            'hostUin': '0',
            'format': 'json',
            'inCharset': 'utf8',
            'outCharset': 'utf-8',
            'notice': '0',
            'platform': 'yqq.json',
            'needNewCode': '0'
        }
        pas = {
            'g_tk_new_20200303': '',
            'g_tk': '',
            'loginUin': '0',
            'hostUin': '0',
            'format': 'json',
            'inCharset': 'utf8',
            'outCharset': 'GB2312',
            'notice': '0',
            'platform': 'yqq.json',
            'needNewCode': '0',
            'cid': '205360772',
            'reqtype': '2',
            'biztype': '1',
            # 歌曲ID
            'topid': ids,
            'cmd': '8',
            'needmusiccrit': '0',
            'pagenum': 0,
            'pagesize': '25',
            'lasthotcommentid': '',
            'domain': 'qq.com',
            'ct': '24',
            'cv': '10101010'
        }
        k += 1
        new_comment_total(pas)
        # print('第%d首评论总数读取成功' % k)
        lyric(par)
        print('第%d首歌词下载成功' % k)
    print('\n歌手%s的歌词下载结束' % song_name_keyword)
    # 最后一定一定要记得保存工作表，否则数据会全没了
    # nb.close()

except FileNotFoundError:
    add_file.append(song_name_keyword)
    print('文件%s-歌曲详细信息.xlsx不存在\n请联系管理员添加' % song_name_keyword)
# print(add_file)
    # 如果文件不存在就将不存在信息保存到CSV文件中
    with open('./lyric/nothingness.csv', 'a+', newline='', encoding='utf-8')as f:
        write = csv.writer(f)
        if add_file:
            write.writerow(add_file)

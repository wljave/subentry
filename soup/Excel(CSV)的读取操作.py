# 引用openpyxl 。
import openpyxl

# **********Excel写入的代码：**************

# 利用openpyxl.Workbook()函数创建新的workbook（工作簿）对象，就是创建新的空的Excel文件。
wb = openpyxl.Workbook()
# wb.active就是获取这个工作簿的活动表，通常就是第一个工作表。
sheet = wb.active
# 可以用.title给工作表重命名。现在第一个工作表的名称就会由原来默认的“sheet1”改为"new title"。
sheet.title = 'new title'
# 把'漫威宇宙'赋值给第一个工作表的A1单元格，就是往A1的单元格中写入了'漫威宇宙'。
sheet['A1'] = '漫威宇宙'
# 先把要写入的多行内容写成列表，再放进大列表里，赋值给rows。
rows = [['美国队长', '钢铁侠', '蜘蛛侠', '雷神'], ['是', '漫威', '宇宙', '经典', '人物']]
# 遍历rows。
for i in rows:
    # 把遍历的内容添加到表格里，这样就实现了多行写入。
    sheet.append(i)
# 打印rows
print(rows)
# 保存新建的Excel文件，并命名为“Marvel.xlsx”
wb.save('Marvel.xlsx')

# **********Excel读取的代码：**************

# 调用openpyxl.load_workbook()函数，打开“Marvel.xlsx”文件。
wb = openpyxl.load_workbook('Marvel.xlsx')
# 获取“Marvel.xlsx”工作簿中名为“new title”的工作表。
sheet = wb['new title']
# sheetnames是用来获取工作簿所有工作表的名字的。如果你不知道工作簿到底有几个工作表，就可以把工作表的名字都打印出来。
sheetname = wb.sheetnames
print(sheetname)
# 把“new title”工作表中A1单元格赋值给A1_cell，再利用单元格value属性，就能打印出A1单元格的值。
A1_cell = sheet['A1']
A1_value = A1_cell.value
print(A1_value)

# ****如果你对openpyxl模块感兴趣，想要有更深入的了解的话，推荐阅读openpyxl模块的官方文档：****
# https://openpyxl.readthedocs.io/en/stable/


# ------------------------------------------------------------------------------------------------------------

# 引用csv模块。
import csv

# **********CSV写入的代码：**************

# 调用open()函数打开csv文件，传入参数：文件名“demo.csv”、写入模式“w”、newline=''、encoding='utf-8'。
csv_file = open('demo.csv', 'w', newline='', encoding='utf-8')
# 用csv.writer()函数创建一个writer对象。
writer = csv.writer(csv_file)
# 调用writer对象的writerow()方法，可以在csv文件里写入一行文字 “电影”和“豆瓣评分”。
writer.writerow(['电影', '豆瓣评分'])
# 在csv文件里写入一行文字 “银河护卫队”和“8.0”。
writer.writerow(['银河护卫队', '8.0'])
# 在csv文件里写入一行文字 “复仇者联盟”和“8.1”。
writer.writerow(['复仇者联盟', '8.1'])
# 写入完成后，关闭文件就大功告成啦！
csv_file.close()

# **********CSV读取的代码：**************


# 用open()打开“demo.csv”文件，'r'是read读取模式，newline=''是避免出现两倍行距。encoding='utf-8'能避免编码问题导致的报错或乱码。
csv_file = open('demo.csv', 'r', newline='', encoding='utf-8')
# 用csv.reader()函数创建一个reader对象。
reader = csv.reader(csv_file)
# 用for循环遍历reader对象的每一行。
for row in reader:
    # 打印row，就能读取出“demo.csv”文件里的内容。
    print(row)
# 读取完成后，要记得关闭文件喔！
csv_file.close()


# csv模块本身还有很多函数和方法，附上csv模块官方文档链接：
# https://yiyibooks.cn/xx/python_352/library/csv.html#module-csv

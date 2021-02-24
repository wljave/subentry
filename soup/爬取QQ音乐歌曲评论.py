import math
import time
import openpyxl
import xlsxwriter
from 使用IP代理模拟浏览器头 import reptile_header

# 初始值
ids_cn = []
song_name_cn = []
comment_totals_cn = []
# A定义ID列    song_id
ids_column = []
# B定义歌曲名列   song_name
song_name_column = []
# C定义歌手列    singer
singer_column = []
# D定义初始化评论总条数   comment_totals
comment_totals = []
# E定义用户昵称列      root_user_nick
root_user_nick_column = []
# F定义用户头像图片链接列      root_avatarurl
root_avatarurl_column = []
# G定义评论时间列      root_comment_time
time_format_column = []
# H定义点赞数列       praise_num
praise_num_column = []

# I定义评论第一回复第一昵称列      reply_nick
one_one_reply_nick_column = []
# J定义评论第一回复第二昵称列      replyed_nick
one_two_reply_nick_column = []
# K定义第一评论第一回复内容列        reply_comment
one_one_comment_content_column = []

# L定义评论第二回复第一昵称列      reply_nick
two_one_reply_nick_column = []
# M定义评论第二回复第二昵称列      replyed_nick
two_two_reply_nick_column = []
# N定义第二评论回复内容列        replyed_content
two_two_comment_content_column = []

# O定义评论内容列      root_comment_content
root_comment_content_column = []


# 定义输入的歌手或歌曲名
song_name_keyword = input('请输入要搜索的歌手或歌曲名(爬取歌曲评论)：')

# # 0打开工作薄
# wb = openpyxl.load_workbook('./crawl_song/%s-歌曲详细信息.xlsx' % song_name_keyword)
# # 1获取工作簿
# # print(wb.active)    # 查看已有的工作簿，并打印
# sheet = wb[song_name_keyword]
# # 2读取A列歌曲ID数据、读取B列歌名数据、读取D列歌手数据
# for cell_A in sheet['A']:
#     # print(cell.value)
#     ids_column.append(cell_A.value)
# for cell_B in sheet['B']:
#     song_name_column.append(cell_B.value)
# for cell_D in sheet['D']:
#     song_column.append(cell_D.value)
# # print(ids_column)
# # 删除A列的表头
# del ids_column[0]
# del song_name_column[0]
# del song_column[0]
# print(ids_column)
# print(song_name_column)
# print(song_column)
#
# 从歌词Excel里面将所有歌曲的评论总数量comment_total取出来，并加入列表comment_totals中
# 0打开工作薄
wily = openpyxl.load_workbook('./lyric/%s-歌词.xlsx' % song_name_keyword)
# 1获取工作簿
sheet_ly = wily[song_name_keyword]
# 2读取A列歌曲ID数据、读取B列歌名数据、读取D列评论总数的数据
for cell_ly_A in sheet_ly['A']:
    ids_cn.append(cell_ly_A.value)
for cell_ly_B in sheet_ly['B']:
    song_name_cn.append(cell_ly_B.value)
for cell_ly_D in sheet_ly['D']:
    comment_totals_cn.append(cell_ly_D.value)
# 删除列的表头
del ids_cn[0]
del song_name_cn[0]
del comment_totals_cn[0]
print(ids_cn)
print(song_name_cn)
print(comment_totals_cn)


# *************将所有歌曲的评论总数量的列表comment_totals保存到CSV文件中********
# with open('./comment/comments.csv', 'a', newline='', encoding='utf-8')as totals:
#     remark = csv.writer(totals)
#     remark.writerow(comment_totals)

# ##############################爬取评论##############################
# 直接引入XHR的JSON链接
def comments(params):
    # 定义最新评论链接
    comment_url = 'https://c.y.qq.com/base/fcgi-bin/fcg_global_comment_h5.fcg'
    # 使用json()方法，将response对象，转为列表/字典
    json_comment = reptile_header(url=comment_url, params=params).json()
    # 查看JSON里面都有什么数据
    # print(json_comment)

    # ******创建Excel表用来保存歌曲评论*******
    # 创建Excel文件
    wbs = xlsxwriter.Workbook('./comment/%s-歌手所有歌曲的评论.xlsx' % song_name_keyword)
    # 创建工作表
    ly_sheet = wbs.add_worksheet(song_name_keyword)
    # 添加工作表样式
    bold = wbs.add_format({
        'bold': True,  # 字体加粗
        'border': 1,  # 单元格边框宽度
        'align': 'left',  # 水平对齐方式
        'valign': 'vcenter',  # 垂直对齐方式
        'fg_color': '#819FF7',  # 单元格背景颜色
        'text_wrap': True,  # 是否自动换行
    })
    # 写入一整行 row:行， col：列， data:要写入的数据, bold:单元格的样式
    ly_sheet.write_row('A1',
                       ['歌曲ID', '歌曲名', '歌手', '评论总数', '用户昵称', '用户头像图片链接', '评论时间', '点赞数', '第一回复第一昵称',
                        '第一回复第二昵称', '第一回复内容', '第二回复第一昵称', '第二回复第二昵称', '第二回复内容', '评论内容'], bold)

    # 一层一层地取字典，获取最新评论列表
    list_comment = json_comment['comment']['commentlist']
    # 循环遍历list_comment
    for comment in list_comment:

        # 最外层获取歌曲ID名列表
        list_song_id = json_comment['topid']
        ids_column.append(int(list_song_id))
        ly_sheet.write_column('A2', ids_column)
        # 最外层获取歌曲名列表
        list_song_name = json_comment['topic_name']
        song_name_column.append(list_song_name)
        ly_sheet.write_column('B2', song_name_column)
        # 歌曲暂时取歌手详细信息里面的记录的是最后一位歌手的名字
        singer_column.append(song_name_keyword)
        ly_sheet.write_column('C2', singer_column)
        # 在comment层获取评论总条数列表
        list_comment_totals = json_comment['comment']['commenttotal']
        comment_totals.append(list_comment_totals)
        ly_sheet.write_column('D2', comment_totals)

        print(
            '--------------------------------------------------------------------------------------------------------')
        # 查找用户昵称
        root_user_nick = comment['nick']
        root_user_nick_column.append(root_user_nick)
        ly_sheet.write_column('E2', root_user_nick_column)
        # 用户头像图片链接
        root_avatarurl = comment['avatarurl']
        root_avatarurl_column.append(root_avatarurl)
        ly_sheet.write_column('F2', root_avatarurl_column)
        # 评论时间
        root_comment_time = comment['time']
        #    # 时间戳转化成时间格式
        times = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(root_comment_time))
        time_format = times[:4] + '年' + times[5:7] + '月' + times[8:10] + '日' + times[10:]
        time_format_column.append(time_format)
        ly_sheet.write_column('G2', time_format_column)
        # 点赞数
        praise_num = comment['praisenum']
        praise_num_column.append(praise_num)
        ly_sheet.write_column('H2', praise_num_column)

        print('【%s】的头像图片链接：' % root_user_nick + root_avatarurl)

        # 判断评论内容是否存在
        try:
            # 若评论内容存在就写入
            root_comment_content = comment['rootcommentcontent']
            root_comment_content_column.append(root_comment_content)
            ly_sheet.write_column('O2', root_comment_content_column)
        except KeyError:
            # 若评论内容键不存在，就写入空值
            root_comment_content = ''
            root_comment_content_column.append(root_comment_content)
            ly_sheet.write_column('O2', root_comment_content_column)

        # 判断回复内容是否为空,若为空将空值写入列表中
        if not comment['middlecommentcontent']:
            # 第一回复第一昵称
            oen_reply_nick = ''
            one_one_reply_nick_column.append(oen_reply_nick)
            ly_sheet.write_column('I2', one_one_reply_nick_column)
            # 第一回复第二昵称
            one_replyed_nick = ''
            one_two_reply_nick_column.append(one_replyed_nick)
            ly_sheet.write_column('J2', one_two_reply_nick_column)
            # 第一回复内容
            one_reply_comment_content = ''
            one_one_comment_content_column.append(one_reply_comment_content)
            ly_sheet.write_column('K2', one_one_comment_content_column)
            # 第二回复第一昵称
            two_reply_nick = ''
            two_one_reply_nick_column.append(two_reply_nick)
            ly_sheet.write_column('L2', two_one_reply_nick_column)
            # 第二回复第二昵称
            two_replyed_nick = ''
            two_two_reply_nick_column.append(two_replyed_nick)
            ly_sheet.write_column('M2', two_two_reply_nick_column)
            # 第二回复内容
            two_replyed_comment_content = ''
            two_two_comment_content_column.append(two_replyed_comment_content)
            ly_sheet.write_column('N2', two_two_comment_content_column)
        else:  # 若回复内容不为空,输出第一条回复
            for middle_comment in comment['middlecommentcontent']:
                # 第一回复第一昵称
                oen_reply_nick = middle_comment['replynick']
                one_one_reply_nick_column.append(oen_reply_nick)
                ly_sheet.write_column('I2', one_one_reply_nick_column)
                # 第一回复第二昵称
                one_replyed_nick = middle_comment['replyednick']
                one_two_reply_nick_column.append(one_replyed_nick)
                ly_sheet.write_column('J2', one_two_reply_nick_column)
                # 第一回复内容
                one_reply_comment_content = middle_comment['subcommentcontent']
                one_one_comment_content_column.append(one_reply_comment_content)
                ly_sheet.write_column('K2', one_one_comment_content_column)
                # 第二回复第一昵称
                two_reply_nick = ''
                two_one_reply_nick_column.append(two_reply_nick)
                ly_sheet.write_column('L2', two_one_reply_nick_column)
                # 第二回复第二昵称
                two_replyed_nick = ''
                two_two_reply_nick_column.append(two_replyed_nick)
                ly_sheet.write_column('M2', two_two_reply_nick_column)
                # 第二回复内容
                two_replyed_comment_content = ''
                two_two_comment_content_column.append(two_replyed_comment_content)
                ly_sheet.write_column('N2', two_two_comment_content_column)
                try:
                    # 判断replyeduin是否为空,若为空输出第二条回复
                    if not middle_comment['replyeduin']:
                        # # 第一回复第一昵称
                        # oen_reply_nick = middle_comment['replynick']
                        # one_one_reply_nick_column.append(oen_reply_nick)
                        # ly_sheet.write_column('I2', one_one_reply_nick_column)
                        # # 第一回复第二昵称
                        # one_replyed_nick = middle_comment['replyednick']
                        # one_two_reply_nick_column.append(one_replyed_nick)
                        # ly_sheet.write_column('J2', one_two_reply_nick_column)
                        # # 第一回复内容
                        # one_reply_comment_content = middle_comment['subcommentcontent']
                        # one_one_comment_content_column.append(one_reply_comment_content)
                        # ly_sheet.write_column('K2', one_one_comment_content_column)
                        # 第二回复第一昵称
                        two_reply_nick = middle_comment['replynick']
                        two_one_reply_nick_column.append(two_reply_nick)
                        ly_sheet.write_column('L2', two_one_reply_nick_column)
                        # 第二回复第二昵称
                        two_replyed_nick = middle_comment['replyednick']
                        two_two_reply_nick_column.append(two_replyed_nick)
                        ly_sheet.write_column('M2', two_two_reply_nick_column)
                        # 第二回复内容
                        two_replyed_comment_content = middle_comment['subcommentcontent']
                        two_two_comment_content_column.append(two_replyed_comment_content)
                        ly_sheet.write_column('N2', two_two_comment_content_column)

                except KeyError:  # 输出第一条回复
                    # # 第一回复昵称
                    # reply_nick = ''
                    # one_reply_nick_column.append(reply_nick)
                    # ly_sheet.write_column('I2', one_reply_nick_column)
                    # # 第二回复昵称
                    # replyed_nick = ''
                    # two_reply_nick_column.append(replyed_nick)
                    # ly_sheet.write_column('J2', two_reply_nick_column)
                    # 第二回复内容
                    two_replyed_comment_content = ''
                    two_two_comment_content_column.append(two_replyed_comment_content)
                    ly_sheet.write_column('L2', two_two_comment_content_column)
    wbs.close()


# 定义每页显示评论的数量（条数）
page_size = 25
it = 0

# 循环遍历页数
for ids in ids_cn:
    # 计算评论总页数=评论总数/每页评论条数       向上取整 math
    pages = math.ceil(comment_totals_cn[it] / page_size)
    # pages = (comment_totals_cn[it] // page_size) + 1
    for pg in range(pages):
        print('\n█████████████【歌手：{}的--{}--歌曲最新评论第{}页】████████████\n'.format(song_name_keyword, song_name_cn[it],
                                                                             pg + 1))
        params = {
            'g_tk_new_20200303': '',
            'g_tk': '',
            'loginUin': '0',
            'hostUin': '0',
            'format': 'json',
            'inCharset': 'utf8',
            'outCharset': 'GB2312',
            'notice': '0',
            'platform': 'yqq.json',
            'needNewCode': '0',
            'cid': '205360772',
            'reqtype': '2',
            'biztype': '1',
            # 歌曲ID
            'topid': ids,
            'cmd': '8',
            'needmusiccrit': '0',
            'pagenum': pg,
            'pagesize': page_size,
            'lasthotcommentid': '',
            'domain': 'qq.com',
            'ct': '24',
            'cv': '10101010'
        }
        pg += 1
        print(pg)
        print(pages)
        if pg == pages:
            comments(params)
            print('没有了，这已经是最后一页了！')
            break
        else:
            comments(params)
    print('\n歌手：%s的歌曲《%s》最新评论下载结束' % (song_name_keyword, song_name_cn[it]))
    it += 1
print('\n歌手%s的所有歌曲的全部评论下载完毕' % song_name_keyword)

# # 如果回复内容为null，则正常评论
# if not comment['middlecommentcontent']:
#     # 用户自己发的的评论
#     root_comment_content = comment['rootcommentcontent']
#     root_comment_content_column.append(root_comment_content)
#     ly_sheet.write_column('N2', root_comment_content_column)
#     print('昵称：' + root_user_nick + '  ━  评论时间：' + time_format + '  ━  点赞量：' + str(
#         praise_num) + '\n' + '评论：' + root_comment_content)
# else:
#     # 用户自己发的的评论
#     root_comment_content = comment['rootcommentcontent']
#     root_comment_content_column.append(root_comment_content)
#     ly_sheet.write_column('N2', root_comment_content_column)
#     for middle_comment in comment['middlecommentcontent']:
#         # 第一回复昵称
#         replyed_nick = middle_comment['replyednick']  # 回复昵称
#         replyed_nick_column.append(replyed_nick)
#         ly_sheet.write_column('I2', replyed_content_column)
#         # 第二回复昵称
#         reply_nick = middle_comment['replynick']
#         reply_nick_column.append(reply_nick)
#         ly_sheet.write_column('K2', reply_nick_column)
#         # 回复评论
#         sub_comment_content = middle_comment['subcommentcontent']
#         # 是否存在回复 @谁：评论：//@谁回复@谁：评论 的情况
#         replyed_uin = middle_comment['replyeduin']
#         if not root_comment_content:
#             print('昵称：' + root_user_nick + '  ━  评论时间：' + time_format + '  ━  点赞量：' + str(
#                 praise_num) + '\n' + '回复 ' + replyed_nick + '：' + sub_comment_content)
#         else:
#             print('昵称：' + root_user_nick + '  ━  评论时间：' + time_format + '  ━  点赞量：' + str(praise_num))
#             print('回复 ' + replyed_nick + '：' + sub_comment_content)
#             print('┝ 评论：%s' % root_comment_content)
#             if not replyed_uin:
#                 print('昵称：' + root_user_nick + '  ━  评论时间：' + time_format + '  ━  点赞量：' + str(praise_num))
#                 print(
#                     '回复 ' + replyed_nick + '：' + sub_comment_content + '//' + reply_nick + '回复 ' + replyed_nick + '：' + sub_comment_content)
#                 print('┝ 评论：%s' % root_comment_content)
# #
#
#
# try:
#     # 回复 @昵称：评论
#     root_comment_nick = comment['rootcommentnick']
#
#     # 用户自己的评论
#     root_comment_content = comment['rootcommentcontent']
#
# except KeyError:
#     # 回复 @昵称：评论
#     root_comment_nick = ''
#     ly_sheet.write_column('J2', root_comment_nick)
#     # 用户自己的评论
#     root_comment_content = ''
#     ly_sheet.write_column('K2', root_comment_content)
# print('【%s】的头像图片链接：' % root_user_nick + root_avatarurl)
#
# if not comment['middlecommentcontent']:
#     # 判断用户昵称是否等于回复昵称，若等于则
#     if root_comment_nick == ('@' + root_user_nick):
#         print('昵称：' + root_user_nick + '  ━  评论时间：' + time_format + '  ━  点赞量：' + str(
#             praise_num) + '\n' + '评论：' + root_comment_content)
# else:
#     middle_comment_content = comment['middlecommentcontent']
#     print('昵称：' + root_user_nick + '  ━  评论时间：' + time_format + '  ━  点赞量：' + str(praise_num))
#     for middle_comment in middle_comment_content:
#         # 回复 @昵称：评论语
#         reply_nick = middle_comment['replyednick']  # 回复昵称
#         ly_sheet.write_column('H2', reply_nick)
#         # 回复评论
#         sub_comment_content = middle_comment['subcommentcontent']
#         ly_sheet.write_column('I2', sub_comment_content)
#         # 当回复昵称不存在时
#         if not reply_nick:
#             print('┝ 评论：回复 %s' % sub_comment_content)
#         else:
#             print('回复 ' + reply_nick + '：' + sub_comment_content)
#             print('┝ 评论：%s' % root_comment_content)


# j = 0
# while j < 60:
#     for ids in ids_column:
#         print('\n█████████████【{}--{}--最新评论第{}页】████████████\n'.format(song_name_keyword, song_name_column[j], i + 1))
#         params = {
#             'g_tk_new_20200303': '',
#             'g_tk': '',
#             'loginUin': '0',
#             'hostUin': '0',
#             'format': 'json',
#             'inCharset': 'utf8',
#             'outCharset': 'GB2312',
#             'notice': '0',
#             'platform': 'yqq.json',
#             'needNewCode': '0',
#             'cid': '205360772',
#             'reqtype': '2',
#             'biztype': '1',
#             # 歌曲ID
#             'topid': ids,
#             'cmd': '8',
#             'needmusiccrit': '0',
#             'pagenum': str(i),
#             'pagesize': '25',
#             'lasthotcommentid': '',
#             'domain': 'qq.com',
#             'ct': '24',
#             'cv': '10101010'
#         }
#         # 循环调用评论方法
#         comments(params)
#     j += 1


# 定义page为每页的索引值
# last_one = True
# # 定义i为歌手名、歌曲名的列表索引
# i = 0
